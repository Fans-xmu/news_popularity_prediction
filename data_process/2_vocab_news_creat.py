# encoding: utf-8
'''
#second process
@author: Fans
@file: 2_vocab_news_creat.py
@time: 2020/12/15 11:54
@desc:change the pkl data[time_list,news_list,label_list] into vocab and news split into [[],[]]
'''
import os
from openpyxl import load_workbook
import jieba
import random
import pickle
from tools.logger import *
import re
import nltk

#读一个文件内的所有news函数----(云杰数据集读法)
def readnews_excel(filename):
    news_list=[]
    workbook = load_workbook(filename)
    sheet = workbook['Sheet1']
    for num in range(2, sheet.max_row):
        news_text = sheet.cell(row=num, column=1).value #依行读取news的内容
        if news_text is None or news_text.strip() == '':
            continue
        if sheet.cell(row=num, column=2).value is None or sheet.cell(row=num, column=3).value is None :
            continue
        sum_comments = int(sheet.cell(row=num, column=2).value)
        positive_comments = int(sheet.cell(row=num, column=3).value)
        #like=int(sheet.cell(row=num, column=4).value)
        #forward=int(sheet.cell(row=num, column=5).value)
        news_list.append([news_text,sum_comments,positive_comments])
    return news_list

def minddataprocess(filename,stop_path):#mind 数据集处理
    stop_words = [w.strip() for w in open(stop_path, 'r', encoding='UTF-8').readlines()]
    punctuations = [',', '.', ':', ';', '?', '(', ')', '[', ']', '&', '!', '*',
                    '@', '#', '$', '%', '\'\'', '\'', '`','``','-', '--', '|', '\/']
    stop_words.extend(punctuations)
    with open(filename, 'rb') as f:
        id_list = pickle.load(f)  #
        time_list = pickle.load(f) #
        news_list = pickle.load(f)  # 所有新闻的content list
        label_positive = pickle.load(f)  # 新闻的热度

    count=0
    news_all = []
    for s in news_list:
        strnew = re.split('[,.?!]', s)
        count+=1
        news_pl = []
        for sent in strnew:
            english=sent.lower()
            str = re.sub('[^\w ]', '', english)
            word_new=nltk.word_tokenize(str)
            #word_new = [w for w in word_new if w not in stop_words and len(w) >= 2]
            news_pl.append(word_new)
        news_all.append(news_pl)
        if count%100==0:
            print("step:",count)
    return news_all,news_list,label_positive,time_list,id_list

def eventdataprocess(filename,stop_path):#灾难数据集
    stop_words = [w.strip() for w in open(stop_path, 'r', encoding='UTF-8').readlines()]
    with open(filename, 'rb') as f:
        time_list = pickle.load(f)  #
        news_list = pickle.load(f)  # 所有新闻的content list
        label_list = pickle.load(f)  # 新闻的热度
    #先分句
    news_all=[]
    for s in news_list:
        str = re.split('[；，。？]',s)
        news_pl=[]
        for sent in str:
            #分词
            word_list=jiebaextract(sent,stop_words)
            news_pl.append(word_list)
        news_all.append(news_pl)

    return news_all,time_list,label_list

def eventdataprocess_seq(filename,stop_path):#灾难数据集
    stop_words = [w.strip() for w in open(stop_path, 'r', encoding='UTF-8').readlines()]
    with open(filename, 'rb') as f:
        time_list = pickle.load(f)  #
        news_list = pickle.load(f)  # 所有新闻的content list
        label_list = pickle.load(f)  # 新闻的热度
    #先分句
    news_all=[]
    for s in news_list:
        word_list=jiebaextract(s,stop_words)
        news_all.append(word_list)
    return news_all,time_list,label_list

#分词函数，对news_list内的每一条新闻进行分词并且去除掉stop_words
def jiebaextract(news_text,stop_words):
    word_list = jieba.cut(news_text)
    #去掉停用词
    word_list = [w for w in word_list if w not in stop_words and len(w) >= 2]
    return word_list

#读取rootdir下所有的excel文件中的新闻以及label
def dataprocess(rootdir,stop_path):
    events = os.listdir(rootdir)
    stop_words = [w.strip() for w in open(stop_path,'r',encoding='UTF-8').readlines()]
    news_list = []
    label_pop = []
    label_positive = []
    word_list = []

    for event in events:
        news_list += readnews_excel(rootdir + event)
    for s in news_list:
        news = s[0]
        wordlist = jiebaextract(news, stop_words)
        word_list.append(wordlist)
        label_pop.append(s[1])
        label_positive.append(s[2])
    return word_list,label_pop,label_positive

#构建词库
def vocab_creat(word_list):
    vocab_dict={}
    #id_dict={}
    count=1
    for wordlist in word_list:
        for word in wordlist:
            for wordq in word:
                if wordq not in vocab_dict.keys():
                    vocab_dict[wordq] = count
                    #id_dict[count]=word
                    count+=1

    vocab_list=list(vocab_dict.keys())
    #对生成的vocab_list进行打乱
    random.shuffle(vocab_list)
    #id打乱
    '''
    for i in range(len(vocab_list)):
        vocab_dict[vocab_list[i]]=i+1
        id_dict[i+1]=vocab_list[i]
    '''
    vocab_num=len(vocab_list)
    return vocab_num,vocab_list

def vocab_creatseq(word_list):
    vocab_dict = {}
    # id_dict={}
    count = 1
    for wordlist in word_list:
        for word in wordlist:
            if word not in vocab_dict.keys():
                vocab_dict[word] = count
                count += 1

    vocab_list = list(vocab_dict.keys())
    # 对生成的vocab_list进行打乱
    random.shuffle(vocab_list)
    # id打乱
    vocab_num = len(vocab_list)
    return vocab_num, vocab_list
if __name__ == "__main__":

    # ---------------------------event chinese dataset process--------------------------------
    filename='../pkldata3/disa_event_data.pkl'
    stop_path = '../pkldata3/停用词词典.txt'

    news_all,time_list,label_list=eventdataprocess(filename,stop_path)
    vocab_num, vocab_list=vocab_creat(news_all)

    #保存所有news的列表以及每个news的label以及时间
    with open('../pkldata/eventnews_data.pkl', 'wb') as f:
        pickle.dump(news_all, f, pickle.HIGHEST_PROTOCOL)
        pickle.dump(label_list, f, pickle.HIGHEST_PROTOCOL)
        pickle.dump(time_list, f, pickle.HIGHEST_PROTOCOL)
    logger.info("[INFO] news_list，label_list,time_list，已保存至pkldata/eventnews_data.pkl文件！")
    #保存vocab_list
    with open('../pkldata/event_vocab.pkl', 'wb') as f:
        pickle.dump(vocab_list, f, pickle.HIGHEST_PROTOCOL)#词库
        pickle.dump(vocab_num, f, pickle.HIGHEST_PROTOCOL)#词的数量
    logger.info("[INFO] vocab_list,vocab_num已保存至pkldata/event_vocab.pkl文件！")

    #--------------------------entertainment dataset处理与上述相同，文件名修改即可-----------

    #---------------------------mind english dataset process--------------------------------
    filename = '../pkldata/mind_data.pkl'
    stop_path = '../pkldata/filter_word.txt'
    news_all, news_list, label_list, time_list, id_list = minddataprocess(filename, stop_path)
    vocab_num, vocab_list = vocab_creat(news_all)
    with open('../pkldata/mindnews_data.pkl', 'wb') as f:
        pickle.dump(news_all, f, pickle.HIGHEST_PROTOCOL)
        pickle.dump(label_list, f, pickle.HIGHEST_PROTOCOL)
        pickle.dump(time_list, f, pickle.HIGHEST_PROTOCOL)
        pickle.dump(id_list, f, pickle.HIGHEST_PROTOCOL)
    logger.info("[INFO] news_list，label_list,time_list have saved to pkldata/mindnews_data.pkl！")

    with open('../pkldata/mind_vocab.pkl', 'wb') as f:
        pickle.dump(vocab_list, f, pickle.HIGHEST_PROTOCOL)
        pickle.dump(vocab_num, f, pickle.HIGHEST_PROTOCOL)
    logger.info("[INFO] vocab_list,vocab_num have saved to pkldata/mind_vocab.pkl！")

